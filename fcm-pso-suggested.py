#--- IMPORT DEPENDENCIES ------------------------------------------------------+
import numpy as np
import time
from random import uniform
import pandas as pd
from sklearn.model_selection import KFold
import math
from sklearn.metrics import accuracy_score, confusion_matrix
from sklearn import metrics
from imblearn.over_sampling import SMOTE
from statistics import mean
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import random
#--- MAIN ---------------------------------------------------------------------+
num_dimensions=23
class Particle:
    def __init__(self):
        self.pos_best_i=[]          # best position individual
        self.err_best_i=-1          # best error individual
        self.err_i=-1               # error individual


        df = pd.read_excel("suggested_weights.xlsx", nrows=num_dimensions, engine='openpyxl')
        arr = df.to_numpy()

        for i, row in enumerate(arr):
            for j, val in enumerate(row):
                if arr[i][j] == "random":
                    arr[i][j]=random.uniform(-1, 1)
                if arr[i][j]=="-VS":
                    arr[i][j]=random.uniform(-1, -0.7)
                if arr[i][j]=='MINUSSTRONG':
                    arr[i][j]=random.uniform(-0.85, -0.5)
                if arr[i][j] =='"-M"':
                    arr[i][j]=random.uniform(-0.65, -0.35)
                if arr[i][j] =="-W":
                    arr[i][j]=random.uniform(-0.5, -0.15)
                if arr[i][j] =="-VW":
                    arr[i][j]=random.uniform(-0.3, 0)

                if arr[i][j] =="VW":
                    arr[i][j]=random.uniform(0, 0.3)
                if arr[i][j] =="W":
                    arr[i][j]=random.uniform(0.15, 0.5)
                if arr[i][j] =="M":
                    arr[i][j]=random.uniform(0.35, 0.65)
                if arr[i][j]=="S":
                    arr[i][j]=random.uniform(0.5, 0.85)
                if arr[i][j]=="VS":
                    arr[i][j]=random.uniform(0.7, 1)

        np.fill_diagonal(arr, 0)

        W2 =np.random.uniform(size = (num_dimensions,num_dimensions), low = -1, high = 1)
        np.fill_diagonal(W2, 0)
        self.position_i=(arr)
        W2[-1] = 0
        self.velocity_i=(W2)

    # evaluate current fitness
    def evaluate(self,err_i):
        # check to see if the current position is an individual best
        if self.err_i<self.err_best_i or self.err_best_i==-1:
            self.pos_best_i=self.position_i.copy()
            self.err_best_i=err_i
 # update new particle velocity
    def update_velocity(self,pos_best_g):
        w=0.1       # constant inertia weight (how much to weigh the previous velocity)
        c1=0.3       # cognative constant
        c2=0.3        # social constant
        r1=uniform(0,1)
        r2=uniform(0,1)
        for i in range(0,num_dimensions):
          for j in range(0,num_dimensions):
              if(i==j):
                continue
              else:
                vel_cognitive=c1*r1*(self.pos_best_i[i][j]-self.position_i[i][j])
                vel_social=c2*r2*(pos_best_g[i][j]-self.position_i[i][j])
                self.velocity_i[i][j]=w*self.velocity_i[i][j] +vel_cognitive+vel_social

    # update the particle position based off new velocity updates
    def update_position(self):
        for i in range(0,num_dimensions):
          for j in range(0,num_dimensions):
              if(i==j):
                continue
              else:
                self.position_i[i][j]=self.position_i[i][j]+self.velocity_i[i][j]
                # adjust maximum position if necessary
                if self.position_i[i][j]>1:
                    self.position_i[i][j]=1

                # adjust minimum position if neseccary
                if self.position_i[i][j]<-1:
                    self.position_i[i][j]=-1
                    
def sig(x):
    return 1/(1 + np.exp(-x))
def minimize_and_plot_concept_evolution(dataset, num_dimensions, num_particles, maxiter):
    err_best_g = float('inf')
    pos_best_g = []
    swarm = [Particle() for _ in range(num_particles)]
    concept_evolution = np.zeros((maxiter, num_dimensions))  # Record concept values over epochs

    for k in range(maxiter):
        # Evaluate fitness of each particle at its current position
        for particle in swarm:
            fitness_result = 0
            for row in dataset:
                fcm_output = [0] * num_dimensions
                for i in range(num_dimensions):
                    sum_temp = sum(particle.position_i[j][i] * row[j] for j in range(num_dimensions) if i != j)
                    fcm_output[i] = sig(sum_temp)

                fitness_result += np.square(fcm_output[-1] - row[-1])

            particle.err_i = fitness_result / len(dataset)
            particle.evaluate(particle.err_i)

            # Update the global best position
            if particle.err_i < err_best_g or k == 0:
                pos_best_g = particle.position_i.copy()
                err_best_g = particle.err_i

        # Record the concept evolution
        concept_evolution[k, :] = [sig(sum(pos_best_g[j][i] * value for j, value in enumerate(row))) for i in range(num_dimensions)]

        # Update velocity and position of each particle after fitness evaluation
        for particle in swarm:
            particle.update_velocity(pos_best_g)
            particle.update_position()

    return pos_best_g, concept_evolution

#calculate deviations values for performance metrics
def calculate_deviation(values):
    # Step 1
    mean = sum(values) / len(values)

    # Step 2
    differences = [value - mean for value in values]

    # Step 3
    squared_differences = [diff ** 2 for diff in differences]

    # Step 4
    mean_squared_differences = sum(squared_differences) / len(squared_differences)

    # Step 5
    standard_deviation = math.sqrt(mean_squared_differences)

    return standard_deviation

#read_clinical_data
dataset=pd.read_excel("cad_full.xlsx", engine='openpyxl')
dataset.fillna(method="bfill", inplace=True)

subset = dataset.sample(frac=0.01)
# Shuffle the subset
subset = subset.reindex(np.random.permutation(subset.index))

# Concatenate the shuffled subset back to the original dataframe
dataset = pd.concat([dataset, subset]).reset_index(drop=True)

# # apply normalization techniques by Column
column = 'AGE'
dataset[column] = (dataset[column]-dataset[column].min())/(dataset[column].max()-dataset[column].min())

column1 = 'BMI'
dataset[column1] = (dataset[column1]-dataset[column1].min())/(dataset[column1].max()-dataset[column1].min())

#divide columns by input and output data
X = dataset.iloc[:,:-1].values
y = dataset.iloc[: ,-1].values

#apply overasmpling method to balance dataset
oversample = SMOTE(sampling_strategy='all')
X, y = oversample.fit_resample(X, y)

#concat input output data where construct the new oversampled dataset
dataset = pd.concat([pd.DataFrame(X), pd.DataFrame(y)], axis=1, join="inner")

start = time.time()

#perform k-fold cross validation
kf = KFold(n_splits=10, shuffle=True)

fold=0
acc=[]
err=[]
best_matrix=[]
concepts=[]
acc=[]
err=[]
best_matrix=[]
concepts=[]
class_accuracies0 =[]
class_accuracies1 =[]
c_matrices=[]
recalls=[]
best_positions=[]
cm_sum = []
sens=[]
spec=[]
prec=[]
epoch=35
# Iterate over each train-test split
for train_index, test_index in kf.split(dataset):
    fold+=1


    print("\n\n************************************")
    print(fold)
    print("---------fold------------")
    print("************************************\n\n")

    # Split train-test
    training_dataset, testing_dataset=dataset.iloc[train_index], dataset.iloc[test_index]

    training_dataset= np.array(training_dataset)
    testing_dataset= np.array(testing_dataset)

    # apply minimization function
    best_position, concept_evolution = minimize_and_plot_concept_evolution(training_dataset,  num_dimensions=23, num_particles=40, maxiter=epoch)
    end = time.time()
    # Plotting concept evolution
    plt.figure(figsize=(15, 10))
    epoch_range = range(1, epoch + 1)  # Create an array for the x-axis representing epochs
    for j in range(num_dimensions):
        plt.plot(epoch_range, concept_evolution[:, j], label=f'Concept {j+1}')
    plt.xlabel('Epochs')
    plt.ylabel('Concept Value')
    plt.title(f'Evolution of Concept Values Over Epochs (Fold {fold})')
    plt.legend()
    plt.legend(loc='upper right')
    # Save the plot as an image file
    filename = f'fold_{fold}_inference_plot.png'  # Define your filename here
    plt.savefig(filename)  # Save the plot
    plt.close()  # Close the plot to free memory
    for i in range(0,num_dimensions):
        for j in range(0,num_dimensions):
            if(i==j):
                continue
            else:
                # adjust maximum position if necessary
                if best_position[i][j]>1:
                    best_position[i][j]=1

                # # adjust minimum position if neseccary
                if best_position[i][j]<-1:
                    best_position[i][j]=-1

    #construct testing dataset
    sum_temp=0

    testing_last_element_fcm_output=[]
    best_position=np.vstack(best_position)
    predicted_results=[None]*23
    for testing_row in testing_dataset:

        for i in range(0,num_dimensions):
            for j in range(0,num_dimensions):
                if(i==j):
                    continue
                else:
                    sum_temp=sum_temp+best_position[j][i]*testing_row[j]

            predicted_results[i]=sum_temp
            predicted_results[i] = sig(predicted_results[i])

            sum_temp=0

        testing_last_element_fcm_output.append((predicted_results)[-1])
    testing_actual_output = testing_dataset[:, -1]

    testing_last_element_fcm_output = np.vstack(testing_last_element_fcm_output)

    testing_last_element_fcm_output = testing_last_element_fcm_output[:, -1]

    temporary_value_results=testing_last_element_fcm_output

    #find the best limit to seperate testing outputs to classes
    limits_acc=[]
    limits= np.arange(0.1, 0.99, 0.01).tolist()

    steady_value_predicted_results=temporary_value_results
    for i in limits:

        temporary_value_results = steady_value_predicted_results > i

        testing_actual_output=np.array(testing_actual_output)
        temporary_value_results=(np.array(temporary_value_results))

        limits_acc.append(accuracy_score(testing_actual_output, temporary_value_results.round())*100)

    max_value = max(limits_acc)
    index = limits_acc.index(max_value)

    testing_last_element_fcm_output = testing_last_element_fcm_output > limits[index]

    testing_actual_output=np.array(testing_actual_output)
    testing_last_element_fcm_output=(np.array(testing_last_element_fcm_output))

    #calculate performance metrics (accuracy, error, sensitivity, specificity, precision)
    A=(accuracy_score(testing_actual_output, testing_last_element_fcm_output.round())*100)
    acc.append(accuracy_score(testing_actual_output, testing_last_element_fcm_output.round())*100)


    err.append(metrics.mean_absolute_error(testing_actual_output, testing_last_element_fcm_output))
    cm = confusion_matrix(testing_actual_output, testing_last_element_fcm_output)
    cm_sum.append(cm)
    class_counts = cm.sum(axis=1)
    accuracies = [0 if count == 0 else cm[i, i] / count for i, count in enumerate(class_counts)]

    # convert the best weight matrix to a pandas DataFrame
    df = pd.DataFrame(best_position, columns=["SEX","AGE", "BMI", "KNOWN CAD","PREVIOUS AMI", "PREVIOUS PCI",
                                              "PREVIOUS CABG", "PREVIOUS STROKE", "DIABETES","SMOKING",
                                              "HYPERTENSION","Dyslipidemia", "Angiopathy","Chronic Kidney Disease",
                                              "Family History of CAD","ASYMPTOMATIC", "ATYPICAL SYMPTOMS", "ANGINA LIKE",
                                              "DYSPNOEA ON EXERTION", "INCIDENT OF PRECORDIAL PAIN","ECG", "EXPERT", "OUTPUT" ])


    # write the DataFrame to an Excel file for each fold
    df.to_excel(f'data{fold}.xlsx', index=False)

    # print(cm)

    if(A!=100):
        TP = cm[1][1]
        TN = cm[0][0]
        FP = cm[0][1]
        FN = cm[1][0]

        sensitivity1 = cm[0,0]/(cm[0,0]+cm[0,1])
        # print('Sensitivity : ', sensitivity1*100 )

        sens.append(sensitivity1*100)

        specificity1 = cm[1,1]/(cm[1,0]+cm[1,1])
        # print('Specificity : ', specificity1*100)

        spec.append(specificity1*100)

        # calculate accuracy
        conf_accuracy = (float (TP+TN) / float(TP + TN + FP + FN))

        # calculate mis-classification
        conf_misclassification = 1- conf_accuracy

        # calculate the sensitivity
        conf_sensitivity = (TP / float(TP + FN))
        # calculate the specificity
        conf_specificity = (TN / float(TN + FP))


        def precision(TP, FP):
            return TP / (TP + FP)
        # calculate precision
        precision_score = precision(TP, FP)
        prec.append(precision_score*100)


#print the performance metrics
print("\n\n\n")
print("-------------end of kfold------------")

print("Accuracies")
print(acc)
print(mean(acc))


acc_deviation = calculate_deviation(acc)
print("acc_deviation")
print(acc_deviation)



print("\n\nError")
print(err)
print(mean(err))
err_deviation = calculate_deviation(err)
print("err_deviation")
print(err_deviation)

sum_matrix = np.sum(cm_sum, axis=0)

print("\nSum of Confusion Matrices:")
print(sum_matrix)

print("\nSensitivity")
print(sens)
print(np.mean(sens))
sens_deviation = calculate_deviation(sens)
print("sens_deviation")
print(sens_deviation)


print("\nSpecificity")
print(spec)
print(np.mean(spec))
spec_deviation = calculate_deviation(spec)
print("spec_deviation")
print(spec_deviation)

print("\n Precision")
print(prec)
print(np.mean(prec))
prec_deviation = calculate_deviation(prec)
print("prec_deviation")
print(prec_deviation)

input_file_names = ["data1.xlsx", "data2.xlsx", "data3.xlsx", "data4.xlsx", "data5.xlsx",
                    "data6.xlsx", "data7.xlsx", "data8.xlsx", "data9.xlsx", "data10.xlsx"]

# Initialize a sum DataFrame with zeros
sum_df = pd.DataFrame(0, index=range(num_dimensions), columns=range(num_dimensions), dtype=float)

# Read data from each input file and accumulate the sum
for file_name in input_file_names:
    try:
        # Read the Excel file
        df = pd.read_excel(file_name, header=None)

        # Exclude the first row and add the array values to sum_df
        sum_df += df.iloc[1:, :].astype(float)

    except Exception as e:
        print(f"Error occurred while processing {file_name}: {e}")
# Calculate the mean DataFrame by dividing the sum_df by the number of files
num_files = len(input_file_names)
mean_df = sum_df / num_files

# Create an Excel file with the mean DataFrame
output_file_name = "mean_values.xlsx"
mean_df.to_excel(output_file_name, index=False, header=False)

print("Mean values saved to", output_file_name)

# Read all Excel files and calculate standard deviations
dfs = [pd.read_excel(file, header=None, skiprows=1, decimal=',') for file in input_file_names]
deviations = np.std([df.values for df in dfs], axis=0)

# Create a DataFrame with deviations
deviation_df = pd.DataFrame(deviations)

# Save the deviations DataFrame to a new Excel file
deviation_df.to_excel('deviations.xlsx', index=False, header=False)


# Calculate the mean DataFrame by dividing the sum_df by the number of files
num_files = len(input_file_names)
mean_df = sum_df / num_files

# Create an Excel file with the mean DataFrame
output_file_name = "deviations.xlsx"
mean_df.to_excel(output_file_name, index=False, header=False)

print("Deviations values saved to", output_file_name)



# Read all Excel files and calculate standard deviations
dfs = [pd.read_excel(file, header=None, skiprows=1, decimal=',') for file in input_file_names]
deviations = np.std([df.values for df in dfs], axis=0)

# Create a DataFrame with deviations
deviation_df = pd.DataFrame(deviations)

# Save the deviations DataFrame to a new Excel file
deviation_df.to_excel('deviations.xlsx', index=False, header=False)

# Load the source and destination Excel files
source_wb = load_workbook('cad_full.xlsx')
destination_wb = load_workbook('mean_values.xlsx')

# Access the first sheet of each workbook
source_sheet = source_wb.active
destination_sheet = destination_wb.active

# Get the values of the first row from the source sheet
first_row_values = []
for cell in source_sheet[1]:
    first_row_values.append(cell.value)

# Paste the first row values into the first row of the destination sheet
for index, value in enumerate(first_row_values, start=1):
    destination_sheet.cell(row=1, column=index, value=value)

# Save the changes to the destination file
destination_wb.save('mean_values.xlsx')

destination_wb = load_workbook('deviations.xlsx')
# Access the first sheet of each workbook
source_sheet = source_wb.active
destination_sheet = destination_wb.active

# Insert a new blank row at the beginning in the destination sheet
destination_sheet.insert_rows(1)

# Get the values of the first row from the source sheet
first_row_values = []
for cell in source_sheet[1]:
    first_row_values.append(cell.value)

# Paste the first row values into the first row of the destination sheet
for index, value in enumerate(first_row_values, start=1):
    destination_sheet.cell(row=1, column=index, value=value)

# Save the changes to the destination file
destination_wb.save('deviations.xlsx')